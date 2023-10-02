import os
import re
import string

import docx2txt
from flask import request, jsonify
from openpyxl import Workbook, load_workbook

from . import teacher
from .. import db
from ..models import EssayCatalog, ParagraphDegree, Paragraph, Essay, ClassInfo, TypeUser, User, Homework, \
    HomeworkResult, HomeworkSentence, EssayResult,Ranking, SentenceResult
import nltk


# nltk.download('punkt')  # 下载必要的数据


@teacher.route('/add_catalog', methods=['POST'])
def add_catalog():
    category_label = request.json.get('category_info')
    textbook_label = request.json.get('textbook_info')
    essayLabel = request.json.get('essay_info')

    if category_label == '':
        return jsonify({'code': 400, 'message': '请填写类别名称'})
    else:
        category_info = EssayCatalog.query.filter_by(label=category_label).first()
        if category_info is None:
            category_id = '00' + str(EssayCatalog.query.filter(EssayCatalog.pid.is_(None)).count() + 1)
            category_catalog = EssayCatalog(id=category_id,
                                            label=category_label,
                                            isEdit='1')
            db.session.add(category_catalog)
            db.session.commit()
        else:
            if textbook_label == '':
                if essayLabel != '':
                    return jsonify({'code': 400, 'message': '请填写书籍名称'})
                return jsonify({'code': 400, 'message': '该类别名称已存在'})
        if textbook_label == '':
            return jsonify({'code': 200})
        else:
            textbook_parent_info = EssayCatalog.query.filter_by(label=category_label).first()
            textbook_parent_id = textbook_parent_info.to_json()['id']
            textbook_info = EssayCatalog.query.filter_by(label=textbook_label, pid=textbook_parent_id).first()
            if textbook_info is None:
                textbook_id = textbook_parent_id + '00' + str(
                    EssayCatalog.query.filter_by(pid=textbook_parent_id).count() + 1)
                textbook_catalog = EssayCatalog(id=textbook_id,
                                                label=textbook_label,
                                                pid=textbook_parent_id,
                                                isEdit='1')
                db.session.add(textbook_catalog)
                db.session.commit()
            else:
                if essayLabel == '':
                    return jsonify({'code': 400, 'message': '该书籍名称已存在'})
            if essayLabel == '':
                return jsonify({'code': 200})
            else:
                essay_parent_info = EssayCatalog.query.filter_by(label=textbook_label, pid=textbook_parent_id).first()
                essay_parent_id = essay_parent_info.to_json()['id']
                essay_info = EssayCatalog.query.filter_by(label=essayLabel, pid=essay_parent_id).first()
                if essay_info is None:
                    essay_id = essay_parent_id + '00' + str(
                        int(EssayCatalog.query.filter_by(pid=essay_parent_id).count()) + 1)

                    essay_catalog = EssayCatalog(id=essay_id,
                                                 label=essayLabel,
                                                 pid=essay_parent_id,
                                                 isEdit='1')
                    db.session.add(essay_catalog)
                    db.session.commit()

                    essay_data = Essay(essay_id=essay_id,
                                       title=essayLabel,
                                       type='empty')
                    db.session.add(essay_data)
                    db.session.commit()

                    return jsonify({'code': 200})
                else:
                    return jsonify({'code': 400, 'message': '该文章名称已存在'})


@teacher.route('/add_all', methods=['POST'])
def add_all():
    essay_id = request.form.get('essay_id')
    essay_file = request.files.get('essay_file')
    audio_file = request.files.get('audio_file')
    whisper_file = request.files.get('whisper_file')
    word_num = request.form.get('word_num')

    # folder = os.path.abspath('..') + r'\dist\assets'
    folder = os.path.abspath('..') + r'\English_vue\public\assets'

    essay_path = os.path.join(folder + r'\files', essay_file.filename)
    essay_file.save(essay_path)

    whisper_path = os.path.join(folder + r'\whisper', whisper_file.filename)
    whisper_file.save(whisper_path)

    audio_path = os.path.join(folder + r'\audio', audio_file.filename)
    audio_file.save(audio_path)

    text = docx2txt.process(essay_file)
    sentences = split_sentences(text, int(word_num))

    whisper_excel = load_workbook(filename=whisper_path)
    sheet = whisper_excel.active
    audio = []
    header = ['text', 'start', 'end']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 将每行数据与表头对应转换为字典
        row_dict = {header[i]: value for i, value in enumerate(row)}
        # 将字典加入数组中
        audio.append(row_dict)

    audio_word = []
    for i in audio:
        current_word = split_punctuation(i['text'])
        for j in range(len(current_word)):
            if j == 0:
                audio_word.append({'word': current_word[j], 'time': i['start']})
            elif j == len(current_word) - 1:
                audio_word.append({'word': current_word[j], 'time': i['end']})
            else:
                word_time = round((float(i['end']) - float(i['start'])) / len(current_word), 2)
                audio_word.append({'word': current_word[j], 'time': round((j * word_time) + float(i['start']), 2)})

    audio_sentence = []
    for i in sentences:
        target_words = split_punctuation(i)

        start = 0
        end = 0
        for j in range(len(audio_word)):
            start_word = target_words[0]
            compare_word = audio_word[j]['word']
            if start_word == compare_word:
                num = 0
                try:
                    for k in range(5):
                        if target_words[k] == audio_word[j + k]['word']:
                            num += 1
                        else:
                            break
                    if num == 5:
                        start = audio_word[j]['time']
                        end = audio_word[j + (len(target_words) - 1)]['time']
                except Exception:
                    pass
        audio_sentence.append({'text': i, 'start': start, 'end': end})

    word_time = round((audio_word[-1]['time'] - audio_word[0]['time']) / len(audio_word), 2)
    for i in range(len(audio_sentence)):

        if audio_sentence[i]['start'] == 0:
            if i == 0:
                audio_sentence[i]['end'] = audio_sentence[i + 1]['start']
                audio_sentence[i]['start'] = audio_sentence[i]['end'] - (
                        len(split_punctuation(audio_sentence[i]['text'])) * word_time)
            elif i == len(audio_sentence) - 1:
                audio_sentence[i]['start'] = audio_sentence[i - 1]['end']
                audio_sentence[i]['end'] = audio_sentence[i]['start'] + (
                        len(split_punctuation(audio_sentence[i]['text'])) * word_time)
            else:
                audio_sentence[i]['start'] = audio_sentence[i - 1]['end']

                if audio_sentence[i + 1]['start'] == 0:
                    audio_sentence[i]['end'] = audio_sentence[i]['start'] + (
                            word_time * len(split_punctuation(audio_sentence[i]['text'])))

                else:
                    audio_sentence[i]['end'] = audio_sentence[i + 1]['start']

    sentences_file = Workbook()
    sentences_worksheet = sentences_file.active
    sentences_worksheet['A1'] = '文章内容'
    sentences_worksheet['B1'] = '音频起始位置'
    sentences_worksheet['C1'] = '音频结束位置'
    for i in audio_sentence:
        sentences_worksheet.append([i['text'], i['start'], i['end']])
    sentences_path = os.path.join(folder + r'\sentences', whisper_file.filename)
    sentences_file.save(sentences_path)

    sen_id = 1
    for i in audio_sentence:
        # translation = Translator(from_lang="en", to_lang="zh").translate(sentences[i])
        # translated_text = translator.translate(sentences[i], dest='zh-cn').text

        paragraphs = Paragraph(essay_id=essay_id,
                               sen_id=sen_id,
                               article=i['text'],
                               audio_start=i['start'],
                               audio_end=i['end'])
        db.session.add(paragraphs)
        db.session.commit()
        sen_id += 1

    select_easy = select(7, sentences)
    easy_id = 1
    for i in select_easy:
        easy = ParagraphDegree(essay_id=essay_id,
                               sen_id=easy_id,
                               word_id=i,
                               grade='简单')
        db.session.add(easy)
        db.session.commit()
        easy_id += 1

    select_med = select(5, sentences)
    med_id = 1
    for i in select_med:
        med = ParagraphDegree(essay_id=essay_id,
                              sen_id=med_id,
                              word_id=i,
                              grade='中等')
        db.session.add(med)
        db.session.commit()
        med_id += 1

    select_hard = select(3, sentences)
    hard_id = 1
    for i in select_hard:
        hard = ParagraphDegree(essay_id=essay_id,
                               sen_id=hard_id,
                               word_id=i,
                               grade='困难')
        db.session.add(hard)
        db.session.commit()
        hard_id += 1

    essay_info = Essay.query.filter_by(essay_id=essay_id).first()
    essay_info.audio_address = '/assets/audio/' + audio_file.filename
    essay_info.essay_address = '/assets/files/' + essay_file.filename
    essay_info.type = 'all'
    db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/add_essay', methods=['POST'])
def add_essay():
    essay_id = request.form.get('essay_id')
    essay_file = request.files.get('essay_file')
    word_num = request.form.get('word_num')

    # folder = os.path.abspath('..') + r'\dist\assets'
    folder = os.path.abspath('..') + r'\English_vue\public\assets'

    essay_path = os.path.join(folder + r'\files', essay_file.filename)
    essay_file.save(essay_path)

    text = docx2txt.process(essay_file)
    sentences = split_sentences(text, int(word_num))

    sen_id = 1
    for i in sentences:
        # translation = Translator(from_lang="en", to_lang="zh").translate(sentences[i])
        # translated_text = translator.translate(sentences[i], dest='zh-cn').text

        paragraphs = Paragraph(essay_id=essay_id,
                               sen_id=sen_id,
                               article=i)
        db.session.add(paragraphs)
        db.session.commit()
        sen_id += 1

    select_easy = select(7, sentences)
    easy_id = 1
    for i in select_easy:
        easy = ParagraphDegree(essay_id=essay_id,
                               sen_id=easy_id,
                               word_id=i,
                               grade='简单')
        db.session.add(easy)
        db.session.commit()
        easy_id += 1

    select_med = select(5, sentences)
    med_id = 1
    for i in select_med:
        med = ParagraphDegree(essay_id=essay_id,
                              sen_id=med_id,
                              word_id=i,
                              grade='中等')
        db.session.add(med)
        db.session.commit()
        med_id += 1

    select_hard = select(3, sentences)
    hard_id = 1
    for i in select_hard:
        hard = ParagraphDegree(essay_id=essay_id,
                               sen_id=hard_id,
                               word_id=i,
                               grade='困难')
        db.session.add(hard)
        db.session.commit()
        hard_id += 1

    essay_info = Essay.query.filter_by(essay_id=essay_id).first()
    essay_info.essay_address = '/assets/files/' + essay_file.filename
    essay_info.type = 'context'
    db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/add_audio', methods=['POST'])
def add_audio():
    essay_id = request.form.get('essay_id')
    audio_file = request.files.get('audio_file')
    whisper_file = request.files.get('whisper_file')

    # folder = os.path.abspath('..') + r'\dist\assets'
    folder = os.path.abspath('..') + r'\English_vue\public\assets'

    whisper_path = os.path.join(folder + r'\whisper', whisper_file.filename)
    whisper_file.save(whisper_path)

    audio_path = os.path.join(folder + r'\audio', audio_file.filename)
    audio_file.save(audio_path)

    whisper_excel = load_workbook(filename=whisper_path)
    sheet = whisper_excel.active
    audio = []
    header = ['text', 'start', 'end']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 将每行数据与表头对应转换为字典
        row_dict = {header[i]: value for i, value in enumerate(row)}
        # 将字典加入数组中
        audio.append(row_dict)

    audio_word = []
    for i in audio:
        current_word = split_punctuation(i['text'])
        for j in range(len(current_word)):
            if j == 0:
                audio_word.append({'word': current_word[j], 'time': i['start']})
            elif j == len(current_word) - 1:
                audio_word.append({'word': current_word[j], 'time': i['end']})
            else:
                word_time = round((float(i['end']) - float(i['start'])) / len(current_word), 2)
                audio_word.append({'word': current_word[j], 'time': round((j * word_time) + float(i['start']), 2)})

    sentences_data = Paragraph.query.filter_by(essay_id=essay_id).all()
    audio_sentence = []
    for sentence in sentences_data:
        article = sentence.to_json()['article']
        target_words = split_punctuation(article)

        start = 0
        end = 0
        for j in range(len(audio_word)):
            start_word = target_words[0]
            compare_word = audio_word[j]['word']
            if start_word == compare_word:
                num = 0
                try:
                    for k in range(5):
                        if target_words[k] == audio_word[j + k]['word']:
                            num += 1
                        else:
                            break
                    if num == 5:
                        start = audio_word[j]['time']
                        end = audio_word[j + (len(target_words) - 1)]['time']
                except Exception:
                    pass
        audio_sentence.append({'text': article, 'start': start, 'end': end})

    word_time = round((audio_word[-1]['time'] - audio_word[0]['time']) / len(audio_word), 2)
    for i in range(len(audio_sentence)):

        if audio_sentence[i]['start'] == 0:
            if i == 0:
                audio_sentence[i]['end'] = audio_sentence[i + 1]['start']
                audio_sentence[i]['start'] = audio_sentence[i]['end'] - (
                        len(split_punctuation(audio_sentence[i]['text'])) * word_time)
            elif i == len(audio_sentence) - 1:
                audio_sentence[i]['start'] = audio_sentence[i - 1]['end']
                audio_sentence[i]['end'] = audio_sentence[i]['start'] + (
                        len(split_punctuation(audio_sentence[i]['text'])) * word_time)
            else:
                audio_sentence[i]['start'] = audio_sentence[i - 1]['end']

                if audio_sentence[i + 1]['start'] == 0:
                    audio_sentence[i]['end'] = audio_sentence[i]['start'] + (
                            word_time * len(split_punctuation(audio_sentence[i]['text'])))

                else:
                    audio_sentence[i]['end'] = audio_sentence[i + 1]['start']

    sentences_file = Workbook()
    sentences_worksheet = sentences_file.active
    sentences_worksheet['A1'] = '文章内容'
    sentences_worksheet['B1'] = '音频起始位置'
    sentences_worksheet['C1'] = '音频结束位置'
    for i in audio_sentence:
        sentences_worksheet.append([i['text'], i['start'], i['end']])
    sentences_path = os.path.join(folder + r'\sentences', whisper_file.filename)
    sentences_file.save(sentences_path)

    for i in range(len(sentences_data)):
        # translation = Translator(from_lang="en", to_lang="zh").translate(sentences[i])
        # translated_text = translator.translate(sentences[i], dest='zh-cn').text

        sentences_data[i].audio_start = audio_sentence[i]['start']
        sentences_data[i].audio_end = audio_sentence[i]['end']
        db.session.commit()

    essay_info = Essay.query.filter_by(essay_id=essay_id).first()
    essay_info.audio_address = '/assets/audio/' + audio_file.filename
    essay_info.type = 'all'
    db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/get_degree', methods=['POST'])
def get_degree():
    essay_id = request.json.get('essay_id')
    data = []
    easy_paragraph = ParagraphDegree.query.filter_by(essay_id=essay_id, grade='简单').first()
    medium_paragraph = ParagraphDegree.query.filter_by(essay_id=essay_id, grade='中等').first()
    hard_paragraph = ParagraphDegree.query.filter_by(essay_id=essay_id, grade='困难').first()
    if easy_paragraph is None:
        data.append({'label': '简单', 'type': '2'})
    else:
        data.append({'label': '简单', 'type': '1'})

    if medium_paragraph is None:
        data.append({'label': '中等', 'type': '2'})
    else:
        data.append({'label': '中等', 'type': '1'})

    if hard_paragraph is None:
        data.append({'label': '困难', 'type': '2'})
    else:
        data.append({'label': '困难', 'type': '1'})

    return jsonify(data)


@teacher.route('/select_list', methods=['POST'])
def select_list():
    essay_id = request.json.get('essay_id')
    grade = request.json.get('grade')
    data = []
    paragraph_info = Paragraph.query.filter_by(essay_id=essay_id).all()

    for i in paragraph_info:
        paragraph = i.to_json()
        sentence_article = paragraph['article']
        word = re.findall(r"[a-zA-Z]+|[-.,!?;:\"\'’”]", sentence_article)

        for index in range(len(word)):
            if index == len(word):
                break
            if word[index] == '\'' and len(word[index + 1]) <= 2 and word[index + 1].isalpha():
                word[index - 1] = word[index - 1] + word[index] + word[index + 1]
                del word[index]
                del word[index]
            if word[index] == '’' and len(word[index + 1]) <= 2 and word[index + 1].isalpha():
                word[index - 1] = word[index - 1] + word[index] + word[index + 1]
                del word[index]
                del word[index]
            if word[index] == '-':
                word[index - 1] = word[index - 1] + word[index] + word[index + 1]
                del word[index]
                del word[index]
            if re.match(r'^[a-z].*[.]$', word[index]):
                new_strings = [word[index][:-1]] + word[index][-1:].split()
                split_index = word.index(word[index])
                word[split_index:split_index + 1] = new_strings
        word_id = 1
        word_list = []

        paragraph_selected = ParagraphDegree.query.filter_by(essay_id=essay_id, sen_id=paragraph['sen_id'],
                                                             grade=grade).first()
        if paragraph_selected is None:
            for j in word:
                if is_english_word(j):
                    word_list.append({'id': word_id, 'text': j, 'type': 'normal'})
                    word_id += 1
                else:
                    word_list.append({'text': j, 'type': 'normal'})
        else:
            selected_id = paragraph_selected.to_json()['word_id'].split(',')
            for j in word:
                if is_english_word(j):
                    if str(word_id) in selected_id:
                        word_list.append({'id': word_id, 'text': j, 'type': 'selected'})
                    else:
                        word_list.append({'id': word_id, 'text': j, 'type': 'normal'})
                    word_id += 1
                else:
                    word_list.append({'text': j})
        data.append({'essay_id': essay_id, 'sen_id': paragraph['sen_id'], 'word_list': word_list})

    return jsonify(data)


@teacher.route('/selected_word', methods=['POST'])
def selected_word():
    selected_list = request.json.get('selected_list')
    for i in selected_list:
        paragraph_degree = ParagraphDegree.query.filter_by(essay_id=i['essay_id'], sen_id=i['sen_id'],
                                                           grade=i['grade']).first()
        if paragraph_degree is None:
            degree_data = ParagraphDegree(essay_id=i['essay_id'],
                                          sen_id=i['sen_id'],
                                          word_id=','.join(str(j) for j in i['select_id']),
                                          grade=i['grade'])
            db.session.add(degree_data)
            db.session.commit()
        else:
            paragraph_degree.essay_id = i['essay_id']
            paragraph_degree.sen_id = i['sen_id']
            paragraph_degree.word_id = ','.join(str(j) for j in i['select_id'])
            paragraph_degree.grade = i['grade']
            db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/sentence_audio', methods=['POST'])
def sentence_audio():
    essay_id = request.json.get('essay_id')
    data = []
    paragraph_info = Paragraph.query.filter_by(essay_id=essay_id).all()
    for i in paragraph_info:
        data.append(i.to_json())

    audio_address = Essay.query.filter_by(essay_id=essay_id).first()
    return jsonify({'data': data, 'audio_address': audio_address.to_json()['audio_address']})


@teacher.route('/save_audio', methods=['POST'])
def save_audio():
    data = request.json.get('data')
    for i in data:
        sentence_info = Paragraph.query.filter_by(essay_id=i['essay_id'], sen_id=i['sen_id']).first()
        sentence_info.audio_start = i['audio_start']
        sentence_info.audio_end = i['audio_end']
        db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/essay_label', methods=['POST'])
def essay_label():
    essay_id = request.json.get('essay_id')
    essay_tile = request.json.get('essay_label')
    essay_info = EssayCatalog.query.filter_by(id=essay_id).first()
    essay_info.label = essay_tile

    essay_info1 = Essay.query.filter_by(essay_id=essay_id).first()
    essay_info1.title = essay_tile
    db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/class_list', methods=['POST'])
def class_list():
    username = request.json.get('username')
    class_list_data = ClassInfo.query.filter_by(teacher_name=username).all()
    data = []
    for i in class_list_data:
        data.append(i.to_json())
    for j in data:
        stu_num = TypeUser.query.filter_by(type_id=j['class_code'], confirm='1').count()
        j['stu_num'] = stu_num
    return jsonify(data)


@teacher.route('/class_student', methods=['POST'])
def class_student():
    class_id = request.json.get('class_id')
    class_data = ClassInfo.query.filter_by(class_id=class_id).first()
    class_list_data = TypeUser.query.filter_by(type_id=class_data.class_code, confirm='1').all()
    data = []
    for i in class_list_data:
        stu_data = User.query.filter_by(user_id=i.to_json()['user_id']).first()
        data.append(stu_data.to_json())
    return jsonify({'data': data, 'class_name': class_data.class_name})


@teacher.route('/class_info', methods=['POST'])
def class_info():
    class_id = request.json.get('class_id')
    class_list_data = ClassInfo.query.filter_by(class_id=class_id).first()
    data = class_list_data.to_json()
    stu_num = TypeUser.query.filter_by(type_id=data['class_code']).count()
    data['stu_num'] = stu_num
    return jsonify(data)


@teacher.route('/add_homework', methods=['POST'])
def add_homework():
    class_id = request.json.get('class_id')
    homework_name = request.json.get('homework_name')
    essay_id = request.json.get('essay_id')
    start_date = request.json.get('start_date')
    end_date = request.json.get('end_date')
    grade = request.json.get('grade')
    homework_type = request.json.get('homework_type')
    print('aa')
    if homework_name == '':
        return jsonify({'code': 400, 'msg': '请输入练习名称'})
    elif essay_id == '':
        return jsonify({'code': 400, 'msg': '请选择练习文章'})
    elif start_date == '':
        return jsonify({'code': 400, 'msg': '请选择开始日期'})
    elif end_date == '':
        return jsonify({'code': 400, 'msg': '请选择结束日期'})
    elif grade == '':
        return jsonify({'code': 400, 'msg': '请选择文章难度'})
    elif homework_type == '':
        return jsonify({'code': 400, 'msg': '请选择练习形式'})
    elif start_date > end_date:
        return jsonify({'code': 400, 'msg': '截止时间不能小于发布时间'})
    else:
        exercise_data = Homework(homework_name=homework_name,
                                 essay_id=essay_id,
                                 class_id=class_id,
                                 start_date=start_date,
                                 end_date=end_date,
                                 grade=grade,
                                 homework_type=homework_type)
        db.session.add(exercise_data)
        db.session.commit()
        return jsonify({'code': 200})


@teacher.route('/class_homework', methods=['POST'])
def class_homework():
    class_id = request.json.get('class_id')
    class_data = Homework.query.filter_by(class_id=class_id).all()
    data = []
    for i in class_data:
        item = i.to_json()
        class_name = ClassInfo.query.filter_by(class_id=item['class_id']).first()
        essay_name = Essay.query.filter_by(essay_id=item['essay_id']).first()
        item['class_name'] = class_name.to_json()['class_name']
        item['essay_name'] = essay_name.to_json()['title']
        item['start_date'] = item['start_date'].strftime("%Y-%m-%d %H:%M:%S")
        item['end_date'] = item['end_date'].strftime("%Y-%m-%d %H:%M:%S")
        data.append(item)

    return jsonify(data)


@teacher.route('/students_homework', methods=['POST'])
def students_homework():
    homework_id = request.json.get('homework_id')
    homework_data = Homework.query.filter_by(homework_id=homework_id).first()
    class_data = ClassInfo.query.filter_by(class_id=homework_data.class_id).first()
    class_student = TypeUser.query.filter_by(type_id=class_data.class_code).all()
    data = []
    for i in class_student:
        student_data = User.query.filter_by(user_id=i.user_id).first()
        student_info = student_data.to_json()
        homework_result = HomeworkResult.query.filter_by(stu_id=i.user_id, homework_id=homework_id).first()
        if homework_result is not None:
            student_info.update(homework_result.to_json())
            student_info['current_time'] = student_info['current_time'].strftime("%Y-%m-%d %H:%M:%S")
            student_info['type'] = '1'
        else:
            student_info['type'] = '0'
        data.append(student_info)

    return jsonify({'data': data, 'class_id': homework_data.to_json()['class_id']})


@teacher.route('/add_class', methods=['POST'])
def add_class():
    class_name = request.json.get('class_name')
    class_code = request.json.get('class_code')
    teacher_name = request.json.get('teacher_name')
    if class_name == '':
        return jsonify({'code': 400, 'msg': '请输入班级名称'})
    else:
        class_data = ClassInfo.query.filter_by(class_name=class_name).first()
        if class_data is not None:
            return jsonify({'code': 400, 'msg': '班级名称已存在'})
    if class_code == '':
        return jsonify({'code': 400, 'msg': '请输入班级码'})
    else:
        class_data = ClassInfo.query.filter_by(class_code=class_code).first()
        if class_data is not None:
            return jsonify({'code': 400, 'msg': '班级码已存在'})

    exercise_data = ClassInfo(class_name=class_name,
                              class_code=class_code,
                              teacher_name=teacher_name)
    db.session.add(exercise_data)
    db.session.commit()
    return jsonify({'code': 200})


@teacher.route('/essay_list', methods=['POST'])
def essay_list():
    user_id = request.json.get('user_id')
    categories = TypeUser.query.filter_by(user_id=user_id).all()
    data = []
    for i in categories:
        category_data = EssayCatalog.query.filter_by(id=i.to_json()['type_id']).first()
        category = category_data.to_json()
        books = EssayCatalog.query.filter_by(pid=category['id']).all()
        for j in books:
            book = j.to_json()
            essays = EssayCatalog.query.filter_by(pid=book['id']).all()
            for k in essays:
                essay = k.to_json()
                essay_data = Essay.query.filter_by(essay_id=essay['id']).first()
                if essay_data.to_json()['type'] != 'empty':
                    data.append(essay_data.to_json())
    return jsonify(data)


@teacher.route('/delete_student', methods=['POST'])
def delete_student():
    class_id = request.json.get('class_id')
    stu_id = request.json.get('stu_id')
    class_stu = TypeUser.query.filter_by(user_id=stu_id, type_id=class_id).first()
    db.session.delete(class_stu)
    homework_stu = HomeworkResult.query.filter_by(stu_id=stu_id).all()
    if len(homework_stu) != 0:
        for i in homework_stu:
            db.session.delete(i)
    homework_sentence = HomeworkSentence.query.filter_by(stu_id=stu_id).all()
    if len(homework_sentence) != 0:
        for j in homework_sentence:
            db.session.delete(j)
    db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/applicant', methods=['POST'])
def applicant():
    class_id = request.json.get('class_id')
    class_data = ClassInfo.query.filter_by(class_id=class_id).first()
    datas = []
    applicant_data = TypeUser.query.filter_by(type_id=class_data.to_json()['class_code'], confirm='0').all()
    for i in applicant_data:
        applicant_info = i.to_json()
        student_data = User.query.filter_by(user_id=applicant_info['user_id']).first()
        student_info = student_data.to_json()
        data = {'user_id': student_info['user_id'], 'username': student_info['username'],
                'avatar': student_info['avatar'], 'confirm': applicant_info['confirm']}
        datas.append(data)
    return jsonify(datas)


@teacher.route('/applicant_confirm', methods=['POST'])
def applicant_confirm():
    user_id = request.json.get('user_id')
    class_id = request.json.get('class_id')
    type = request.json.get('type')
    class_data = ClassInfo.query.filter_by(class_id=class_id).first()
    student_data = TypeUser.query.filter_by(user_id=user_id, type_id=class_data.to_json()['class_code']).first()
    student_data.confirm = type
    db.session.commit()

    return jsonify({'code': 200})


@teacher.route('/get_catalog', methods=['POST'])
def get_catalog():
    user_id = request.json.get('user_id')
    user_catalog = TypeUser.query.filter_by(user_id=user_id).all()
    data = []
    for i in user_catalog:
        catalog_data = i.to_json()
        catalog = EssayCatalog.query.filter_by(id=catalog_data['type_id']).first()
        catalog_info = catalog.to_json()
        data.append(catalog_info['label'])
    return jsonify(data)


@teacher.route('/delete_data', methods=['POST'])
def delete_data():
    essay_id = request.json.get('essay_id')
    essay_data = Essay.query.filter_by(essay_id=essay_id).first()
    essay_data.essay_address = ''
    essay_data.audio_address = ''
    essay_data.type = 'empty'
    paragraphs = Paragraph.query.filter_by(essay_id=essay_id).all()
    for a in paragraphs:
        db.session.delete(a)
    paragraph_degree = ParagraphDegree.query.filter_by(essay_id=essay_id).all()
    for b in paragraph_degree:
        db.session.delete(b)
    essay_result = EssayResult.query.filter_by(essay_id=essay_id).all()
    for c in essay_result:
        user_data = User.query.filter_by(user_id=c.user_id).first()
        ranking_data = Ranking.query.filter_by(user_name=user_data.username).first()
        if ranking_data.essay_num == '0':
            db.session.delete(ranking_data)
        else:
            score = int(ranking_data.essay_num)*int(ranking_data.score) - int(c.score)
            essay_num = int(ranking_data.essay_num)-1
            ranking_data.essay_num = str(essay_num)
            ranking_data.score = round(score / essay_num, 0)

        db.session.delete(c)
    sentence_result = SentenceResult.query.filter_by(essay_id=essay_id).all()
    for g in sentence_result:
        db.session.delete(g)
    homework_data = Homework.query.filter_by(essay_id=essay_id).all()
    for d in homework_data:
        homework_result = HomeworkResult.query.filter_by(homework_id=d.homework_id).all()
        for e in homework_result:
            db.session.delete(e)
        homework_sentence = HomeworkSentence.query.filter_by(homework_id=d.homework_id).all()
        for f in homework_sentence:
            db.session.delete(f)
        db.session.delete(d)

    return jsonify({'code': 200})


def split_sentences(text, word_num):
    sentences = nltk.sent_tokenize(text)
    new_sentences = []
    sub_sentence = ""

    for sentence in sentences:
        words = sentence.split()
        if len(words) < word_num:
            sub_word = sub_sentence.split()
            if len(sub_word) + len(words) > word_num:
                new_sentences.append(sub_sentence.strip())
                sub_sentence = sentence
            else:
                sub_sentence = sub_sentence + " " + sentence
        else:
            if len(sub_sentence) != 0:
                new_sentences.append(sub_sentence.strip())
                sub_sentence = ""
            new_sentences.append(sentence.strip())
    if sub_sentence != "":
        new_sentences.append(sub_sentence.strip())

    return new_sentences


def is_english_word(word):
    pattern = re.compile('^[a-zA-Z].*[a-zA-Z]$|^[a-zA-Z]$|^[A-Z].*[.]$')
    return bool(pattern.match(word))


def split_punctuation(sentence):
    words = sentence.split()
    punctuation_string = string.punctuation
    # 删掉标点符号
    for word in range(len(words)):
        for j in punctuation_string:
            words[word] = words[word].replace(j, '').lower()
    word_list = [word for word in words if any(char.isalpha() for char in word)]
    return word_list


def split_sentence(sentence):
    word = re.findall(r"[a-zA-Z]+|[-.,!?;:\"\'’”]", sentence)

    for index in range(len(word)):
        if index == len(word):
            break
        if word[index] == '\'' and len(word[index + 1]) <= 2 and word[index + 1].isalpha():
            word[index - 1] = word[index - 1] + word[index] + word[index + 1]
            del word[index]
            del word[index]
        if word[index] == '’' and len(word[index + 1]) <= 2 and word[index + 1].isalpha():
            word[index - 1] = word[index - 1] + word[index] + word[index + 1]
            del word[index]
            del word[index]
        if word[index] == '-':
            word[index - 1] = word[index - 1] + word[index] + word[index + 1]
            del word[index]
            del word[index]
        if re.match(r'^[a-z].*[.]$', word[index]):
            new_strings = [word[index][:-1]] + word[index][-1:].split()
            split_index = word.index(word[index])
            word[split_index:split_index + 1] = new_strings
    return word


def select(group_size, sentences):
    select_id = []
    for sentence in sentences:
        words = split_sentence(sentence)
        index = 1
        temp = []
        temp_id = []
        temp_word = []
        for word in words:
            if is_english_word(word):
                temp.append({'id': index, 'word': word})
                index += 1

        num_groups = len(temp) // group_size  # 计算组的数量

        if len(temp) % group_size != 0:
            num_groups += 1
        for j in range(num_groups):
            start_index = j * group_size
            end_index = start_index + group_size
            group = temp[start_index:end_index]

            max_num = 0
            max_word = ''
            max_index = ''
            for k in group:
                if len(k['word']) > max_num:
                    if k['word'] not in temp_word:
                        max_num = len(k['word'])
                        max_word = k['word']
                        max_index = str(k['id'])
            temp_id.append(max_index)
            temp_word.append(max_word)

        id_str = ','.join(temp_id)
        select_id.append(id_str)
    return select_id
