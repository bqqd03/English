import string

import docx2txt
import nltk
from openpyxl.reader.excel import load_workbook


def split_punctuation(aa):
    bb = aa.split()
    punctuation_string = string.punctuation
    # 删掉标点符号
    for word in range(len(bb)):
        for j in punctuation_string:
            bb[word] = bb[word].replace(j, '').lower()
    word_list = [word for word in bb if any(char.isalpha() for char in word)]
    return word_list


text = docx2txt.process(r'D:\英语练习资源\1-Unit 3 Text A\1-Unit 3 Text A.docx')
sentences = nltk.sent_tokenize(text)
new_sentences = []
sub_sentence = ""

for sentence in sentences:
    words = sentence.split()
    if len(words) < 40:
        sub_word = sub_sentence.split()
        if len(sub_word) + len(words) > 40:
            new_sentences.append(sub_sentence.strip())
            sub_sentence = sentence
        else:
            sub_sentence = sub_sentence + " " + sentence
    else:
        if len(sub_sentence) != 0:
            new_sentences.append(sub_sentence.strip())
            sub_sentence = ""
        new_sentences.append(sentence.strip())
if sub_sentence != "":
    new_sentences.append(sub_sentence.strip())

whisper_excel = load_workbook(filename=r'D:\英语练习资源\1-Unit 3 Text A\1-Unit 3 Text A.xlsx')
sheet = whisper_excel.active
audio = []
header = ['text', 'start', 'end']
for row in sheet.iter_rows(min_row=2, values_only=True):
    # 将每行数据与表头对应转换为字典
    row_dict = {header[i]: value for i, value in enumerate(row)}
    # 将字典加入数组中
    audio.append(row_dict)

audio_word = []
for i in audio:
    current_word = split_punctuation(i['text'])
    for j in range(len(current_word)):
        if j == 0:
            audio_word.append({'word': current_word[j], 'time': i['start']})
        elif j == len(current_word) - 1:
            audio_word.append({'word': current_word[j], 'time': i['end']})
        else:
            word_time = round((float(i['end']) - float(i['start'])) / len(current_word), 2)
            audio_word.append({'word': current_word[j], 'time': round((j * word_time) + float(i['start']), 2)})

audio_sentence = []
for i in sentences:
    target_words = split_punctuation(i)

    start = 0
    end = 0
    for j in range(len(audio_word)):
        start_word = target_words[0]
        compare_word = audio_word[j]['word']
        if start_word == compare_word:
            num = 0
            try:
                for k in range(5):
                    if target_words[k] == audio_word[j + k]['word']:
                        num += 1
                    else:
                        break
                if num == 5:
                    start = audio_word[j]['time']
                    end = audio_word[j + (len(target_words) - 1)]['time']
            except Exception:
                pass
    audio_sentence.append({'text': i, 'start': start, 'end': end})
for i in audio_sentence:
    print(i)
